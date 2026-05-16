import os
from datetime import date
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import select, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.member import Member
from app.models.indicator import IndicatorData
from app.models.report import Report
from app.models.health_event import HealthEvent
from app.models.hospital import HospitalEvent
from app.models.vaccine import VaccineRecord
from app.models.reminder import Reminder


# Status color mapping
STATUS_COLORS = {
    "normal": "C6F6D5",
    "low": "FEF3C7",
    "high": "FEF3C7",
    "critical": "FECACA",
    "pending": "E2E8F0",
    "completed": "C6F6D5",
    "overdue": "FECACA",
}

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def _auto_width(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _add_header_row(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _add_data_row(ws, row_num, values, status_key=None, status_idx=None):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")
        if status_idx is not None and col == status_idx + 1:
            color = STATUS_COLORS.get(str(value), "FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


async def export_excel(
    db: AsyncSession,
    member: Member,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> BytesIO:
    wb = Workbook()

    # Sheet 1: Member Info
    ws_info = wb.active
    ws_info.title = "成员信息"
    ws_info.append(["姓名", member.name])
    ws_info.append(["类型", member.type])
    ws_info.append(["性别", member.gender])
    ws_info.append(["血型", member.blood_type or "未填写"])
    ws_info.append(["出生日期", str(member.birth_date) if member.birth_date else "未填写"])
    ws_info.append(["导出日期", str(date.today())])
    _auto_width(ws_info)

    # Sheet 2: Indicators
    ws_ind = wb.create_sheet("健康指标")
    headers = ["指标名称", "数值", "单位", "状态", "记录日期"]
    _add_header_row(ws_ind, headers)

    stmt = (
        select(IndicatorData)
        .where(IndicatorData.member_id == member.id)
        .order_by(desc(IndicatorData.record_date))
    )
    if start_date:
        stmt = stmt.where(IndicatorData.record_date >= start_date)
    if end_date:
        stmt = stmt.where(IndicatorData.record_date <= end_date)
    result = await db.execute(stmt)
    indicators = result.scalars().all()

    for idx, ind in enumerate(indicators, 2):
        status_label = {"normal": "正常", "low": "偏低", "high": "偏高", "critical": "严重异常"}.get(ind.status, ind.status)
        _add_data_row(ws_ind, idx, [ind.indicator_name, float(ind.value), ind.unit, status_label, str(ind.record_date)], status_key=ind.status, status_idx=3)
    _auto_width(ws_ind)

    # Sheet 3: Reports
    ws_rep = wb.create_sheet("检查报告")
    headers = ["类型", "医院", "日期", "OCR状态"]
    _add_header_row(ws_rep, headers)

    stmt = select(Report).where(Report.member_id == member.id).order_by(desc(Report.report_date))
    result = await db.execute(stmt)
    reports = result.scalars().all()
    for idx, r in enumerate(reports, 2):
        type_label = {"lab": "检验", "diagnosis": "诊断", "prescription": "处方", "discharge": "出院"}.get(r.type, r.type)
        ocr_label = {"pending": "待处理", "processing": "识别中", "completed": "已完成", "failed": "失败"}.get(r.ocr_status, r.ocr_status)
        _add_data_row(ws_rep, idx, [type_label, r.hospital or "未填写", str(r.report_date) if r.report_date else "未填写", ocr_label])
    _auto_width(ws_rep)

    # Sheet 4: Health Events
    ws_evt = wb.create_sheet("健康事件")
    headers = ["类型", "日期", "医院", "诊断"]
    _add_header_row(ws_evt, headers)

    stmt = select(HealthEvent).where(HealthEvent.member_id == member.id).order_by(desc(HealthEvent.event_date))
    result = await db.execute(stmt)
    events = result.scalars().all()
    type_map = {"visit": "就诊", "lab": "检验", "medication": "用药", "symptom": "症状", "hospital": "住院", "vaccine": "疫苗", "checkup": "体检", "milestone": "里程碑"}
    for idx, evt in enumerate(events, 2):
        _add_data_row(ws_evt, idx, [type_map.get(evt.type, evt.type), str(evt.event_date), evt.hospital or "", evt.diagnosis or ""])
    _auto_width(ws_evt)

    # Sheet 5: Hospital Events
    ws_hos = wb.create_sheet("住院记录")
    headers = ["医院", "科室", "入院日期", "出院日期", "诊断", "状态"]
    _add_header_row(ws_hos, headers)

    stmt = select(HospitalEvent).where(HospitalEvent.member_id == member.id).order_by(desc(HospitalEvent.admission_date))
    result = await db.execute(stmt)
    hospitals = result.scalars().all()
    for idx, h in enumerate(hospitals, 2):
        status_label = "住院中" if h.status == "active" else "已出院"
        _add_data_row(ws_hos, idx, [h.hospital, h.department or "", str(h.admission_date), str(h.discharge_date) if h.discharge_date else "未出院", h.diagnosis or "", status_label])
    _auto_width(ws_hos)

    # Sheet 6: Vaccines
    ws_vac = wb.create_sheet("疫苗记录")
    headers = ["疫苗名称", "剂次", "计划日期", "接种日期", "状态", "地点"]
    _add_header_row(ws_vac, headers)

    stmt = select(VaccineRecord).where(VaccineRecord.member_id == member.id).order_by(desc(VaccineRecord.scheduled_date))
    result = await db.execute(stmt)
    vaccines = result.scalars().all()
    status_map = {"completed": "已完成", "pending": "待接种", "upcoming": "即将接种", "overdue": "已逾期"}
    for idx, v in enumerate(vaccines, 2):
        _add_data_row(ws_vac, idx, [v.vaccine_name, v.dose or 1, str(v.scheduled_date) if v.scheduled_date else "", str(v.actual_date) if v.actual_date else "未接种", status_map.get(v.status, v.status), v.location or ""])
    _auto_width(ws_vac)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def export_pdf(
    db: AsyncSession,
    member: Member,
) -> BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    # Try to use a Chinese font
    chinese_style = ParagraphStyle(
        'Chinese',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
    )
    title_style = ParagraphStyle(
        'ChineseTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        alignment=1,  # center
    )
    heading_style = ParagraphStyle(
        'ChineseHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
    )

    story = []

    # Title
    story.append(Paragraph(f"{member.name} - Health Report", title_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Generated: {date.today()}", chinese_style))
    story.append(Spacer(1, 1*cm))

    # Basic Info
    story.append(Paragraph("Member Information", heading_style))
    info_data = [
        ["Name", member.name],
        ["Type", member.type],
        ["Gender", member.gender],
        ["Blood Type", member.blood_type or "N/A"],
    ]
    info_table = Table(info_data, colWidths=[4*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor("#1E293B")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#E2E8F0")),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 1*cm))

    # Recent Indicators
    story.append(Paragraph("Recent Indicators", heading_style))
    stmt = (
        select(IndicatorData)
        .where(IndicatorData.member_id == member.id)
        .order_by(desc(IndicatorData.record_date))
        .limit(10)
    )
    result = await db.execute(stmt)
    indicators = result.scalars().all()

    if indicators:
        ind_data = [["Indicator", "Value", "Unit", "Status", "Date"]]
        for ind in indicators:
            status_color = {"normal": colors.HexColor("#10B981"), "low": colors.HexColor("#F59E0B"), "high": colors.HexColor("#F59E0B"), "critical": colors.HexColor("#EF4444")}.get(ind.status, colors.black)
            ind_data.append([ind.indicator_name, str(float(ind.value)), ind.unit, ind.status, str(ind.record_date)])
        ind_table = Table(ind_data, colWidths=[4*cm, 3*cm, 2.5*cm, 2.5*cm, 3*cm])
        ind_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(ind_table)
    else:
        story.append(Paragraph("No indicator records found.", chinese_style))

    story.append(Spacer(1, 1*cm))

    # Recent Reports
    story.append(Paragraph("Recent Reports", heading_style))
    stmt = select(Report).where(Report.member_id == member.id).order_by(desc(Report.report_date)).limit(5)
    result = await db.execute(stmt)
    reports = result.scalars().all()

    if reports:
        rep_data = [["Type", "Hospital", "Date", "OCR Status"]]
        for r in reports:
            rep_data.append([r.type, r.hospital or "N/A", str(r.report_date) if r.report_date else "N/A", r.ocr_status])
        rep_table = Table(rep_data, colWidths=[3*cm, 5*cm, 3*cm, 3*cm])
        rep_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(rep_table)
    else:
        story.append(Paragraph("No reports found.", chinese_style))

    # Disclaimer
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph(
        "Disclaimer: This report is for reference only and does not replace professional medical advice.",
        ParagraphStyle('Disclaimer', parent=chinese_style, fontSize=8, textColor=colors.grey)
    ))

    doc.build(story)
    output.seek(0)
    return output
